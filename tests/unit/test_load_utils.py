import os
import sys
import uuid
from datetime import date, timedelta

import pandas as pd
import psycopg2
import pytest
from openpyxl import Workbook

# Добавляем путь к scripts в sys.path, чтобы импортировать load_utils
sys.path.append(os.path.join(os.path.dirname(__file__), "..", ".."))

from scripts.load_utils import (
    _canonicalize_headers,
    _get_discount_from_cell,
    _norm,
    _norm_key,
    _to_float,
    _to_int,
    get_conn,
    read_any,
    upsert_records,
)

# -----------------------------------------------------------------------------
# Управление интеграционными DB-тестами
# -----------------------------------------------------------------------------
# По умолчанию пропускаем тесты, которые требуют живую PostgreSQL.
# Включить можно, задав RUN_DB_TESTS=1 (или true/yes) в окружении.
RUN_DB_TESTS = os.getenv("RUN_DB_TESTS", "0").lower() in ("1", "true", "yes")


# =============================================================================
# Tests for _norm() function
# =============================================================================


@pytest.mark.unit
def test_norm_removes_leading_and_trailing_spaces():
    """
    Test: _norm() should remove spaces at the beginning and end.
    """
    input_string = "  hello world  "
    result = _norm(input_string)
    assert result == "hello world"


@pytest.mark.unit
def test_norm_replaces_newlines_with_spaces():
    """
    Test: _norm() should replace \\n with space.
    """
    input_string = "hello\nworld"
    result = _norm(input_string)
    assert result == "hello world"


@pytest.mark.unit
def test_norm_handles_none():
    """
    Test: _norm(None) should return empty string.
    """
    input_value = None
    result = _norm(input_value)
    assert result == ""


@pytest.mark.unit
def test_norm_collapses_multiple_spaces():
    """
    Test: _norm() should collapse multiple spaces into one.
    """
    input_string = "hello    world"
    result = _norm(input_string)
    assert result == "hello world"


# =============================================================================
# Tests for _to_float() function
# =============================================================================


@pytest.mark.unit
def test_to_float_converts_string_to_float():
    """
    Test: _to_float() should convert string "1000" to 1000.0
    """
    input_string = "1000"
    result = _to_float(input_string)
    assert result == 1000.0


@pytest.mark.unit
def test_to_float_handles_spaces_in_numbers():
    """
    Test: _to_float() should handle "1 000" (with spaces).
    """
    input_string = "1 000"
    result = _to_float(input_string)
    assert result == 1000.0


@pytest.mark.unit
def test_to_float_handles_comma_as_decimal_separator():
    """
    Test: _to_float() should convert "1000,50" to 1000.5
    """
    input_string = "1000,50"
    result = _to_float(input_string)
    assert result == 1000.5


@pytest.mark.unit
def test_to_float_handles_none():
    """
    Test: _to_float(None) should return None.
    """
    input_value = None
    result = _to_float(input_value)
    assert result is None


@pytest.mark.unit
def test_to_float_handles_negative_numbers():
    """
    Test: _to_float() should handle negative numbers.
    """
    input_string = "-42.5"
    result = _to_float(input_string)
    assert result == -42.5


# =============================================================================
# Tests for _to_int() function
# =============================================================================


@pytest.mark.unit
def test_to_int_converts_string_to_int():
    """
    Test: _to_int() should convert "42" to 42.
    """
    input_string = "42"
    result = _to_int(input_string)
    assert result == 42


@pytest.mark.unit
def test_to_int_rounds_float_to_int():
    """
    Test: _to_int() should round "42.7" to 43.
    """
    input_string = "42.7"
    result = _to_int(input_string)
    assert result == 43


@pytest.mark.unit
def test_to_int_handles_none():
    """
    Test: _to_int(None) should return None.
    """
    input_value = None
    result = _to_int(input_value)
    assert result is None


# =============================================================================
# Tests for _norm_key() function
# =============================================================================


@pytest.mark.unit
def test_norm_key_converts_to_lowercase():
    """
    Test: _norm_key() should convert to lowercase.
    """
    input_string = "ЦЕНА"
    result = _norm_key(input_string)
    assert result == "цена"


@pytest.mark.unit
def test_norm_key_replaces_spaces_with_underscores():
    """
    Test: _norm_key() should replace spaces with underscores.
    """
    input_string = "Цена прайс"
    result = _norm_key(input_string)
    assert result == "цена_прайс"


@pytest.mark.unit
def test_norm_key_replaces_yo_with_ye():
    """
    Test: _norm_key() should replace 'ё' with 'е'.
    """
    input_string = "объём"
    result = _norm_key(input_string)
    assert result == "объем"


@pytest.mark.unit
def test_norm_key_handles_percent_sign():
    """
    Test: _norm_key() removes special characters including percent sign.
    """
    input_string = "Алк, %"
    result = _norm_key(input_string)
    assert result == "алк"


@pytest.mark.unit
def test_norm_key_removes_special_characters():
    """
    Test: _norm_key() should remove special characters (commas, dots, etc).
    """
    input_string = "Цена, руб."
    result = _norm_key(input_string)
    assert result == "цена_руб"


@pytest.mark.unit
def test_norm_key_collapses_multiple_underscores():
    """
    Test: _norm_key() should collapse multiple underscores into one.
    """
    input_string = "Цена___прайс"
    result = _norm_key(input_string)
    assert result == "цена_прайс"


# =============================================================================
# Tests for _get_discount_from_cell() function
# =============================================================================


class TestGetDiscountFromCell:
    """
    Тесты для извлечения скидки из ячейки Excel.

    Функция _get_discount_from_cell() читает значение из указанной ячейки
    Excel файла и парсит скидку в формате доли (0.0 - 1.0).
    """

    @pytest.mark.unit
    def test_discount_percentage_format(self, tmp_path):
        """
        Test: Скидка в формате "10%" должна преобразоваться в 0.10
        """
        wb = Workbook()
        ws = wb.active
        ws["S5"] = "10%"
        excel_file = tmp_path / "test_discount.xlsx"
        wb.save(excel_file)

        result = _get_discount_from_cell(str(excel_file), 0, "S5")
        assert result == 0.10

    @pytest.mark.unit
    def test_discount_decimal_format(self, tmp_path):
        """
        Test: Скидка в формате 0.15 (десятичная дробь) должна остаться 0.15
        """
        wb = Workbook()
        ws = wb.active
        ws["S5"] = 0.15
        excel_file = tmp_path / "test_discount_decimal.xlsx"
        wb.save(excel_file)

        result = _get_discount_from_cell(str(excel_file), 0, "S5")
        assert result == 0.15

    @pytest.mark.unit
    def test_discount_large_number_format(self, tmp_path):
        """
        Test: Скидка 10 (число > 1) должна преобразоваться в 0.10
        """
        wb = Workbook()
        ws = wb.active
        ws["S5"] = 10
        excel_file = tmp_path / "test_discount_large.xlsx"
        wb.save(excel_file)

        result = _get_discount_from_cell(str(excel_file), 0, "S5")
        assert result == 0.10

    @pytest.mark.unit
    def test_discount_empty_cell(self, tmp_path):
        """
        Test: Пустая ячейка S5 должна возвращать None
        """
        wb = Workbook()
        wb.active  # S5 не трогаем
        excel_file = tmp_path / "test_discount_empty.xlsx"
        wb.save(excel_file)

        result = _get_discount_from_cell(str(excel_file), 0, "S5")
        assert result is None

    @pytest.mark.unit
    def test_discount_invalid_format(self, tmp_path):
        """
        Test: Некорректный формат ("invalid") должен возвращать None
        """
        wb = Workbook()
        ws = wb.active
        ws["S5"] = "invalid"
        excel_file = tmp_path / "test_discount_invalid.xlsx"
        wb.save(excel_file)

        result = _get_discount_from_cell(str(excel_file), 0, "S5")
        assert result is None


# =============================================================================
# High-level helpers
# =============================================================================


class TestCanonicalizeHeaders:
    """
    Тесты для функции _canonicalize_headers()
    """

    @pytest.mark.unit
    def test_canonicalize_headers_maps_correctly(self):
        cols = ["Код", "Производитель", "Название", "Цена прайс", "Цена со скидкой"]
        result = _canonicalize_headers(cols)
        assert result["Код"] == "code"
        assert result["Производитель"] == "producer"
        assert result["Название"] == "title_ru"
        assert result["Цена прайс"] == "price_rub"
        assert result["Цена со скидкой"] == "price_discount"

    @pytest.mark.unit
    def test_canonicalize_headers_handles_unnamed(self):
        cols = ["Unnamed: 0", "Код", "", "Цена", "unnamed_5"]
        result = _canonicalize_headers(cols)
        assert result["Unnamed: 0"] is None
        assert result["Код"] == "code"
        assert result[""] is None
        assert result["Цена"] == "price_rub"
        assert result["unnamed_5"] is None

    @pytest.mark.unit
    def test_canonicalize_headers_handles_discount_with_percent(self):
        cols = ["Код", "Цена со скидкой 10%"]
        result = _canonicalize_headers(cols)
        assert result["Код"] == "code"
        assert result["Цена со скидкой 10%"] == "price_discount"


class TestReadAny:
    """
    Тесты для функции read_any()
    """

    @pytest.mark.unit
    def test_read_any_detects_csv_encoding(self, tmp_path):
        # CSV с русскими символами в cp1251
        csv_file = tmp_path / "test_encoding.csv"
        csv_content = "Код,Название,Цена\n123,Тестовое вино,1000\n456,Другое вино,2000"
        csv_file.write_bytes(csv_content.encode("cp1251"))

        df = read_any(str(csv_file))

        assert "code" in df.columns
        assert len(df) == 2
        assert df.iloc[0]["code"] == "123"

    @pytest.mark.unit
    def test_read_any_handles_excel_basic(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Код"
        ws["B1"] = "Название"
        ws["C1"] = "Цена"
        ws["A2"] = "123"
        ws["B2"] = "Вино"
        ws["C2"] = "1000"

        excel_file = tmp_path / "test_basic.xlsx"
        wb.save(excel_file)

        df = read_any(str(excel_file))

        assert "code" in df.columns
        assert "title_ru" in df.columns
        assert "price_rub" in df.columns
        assert len(df) == 1
        assert df.iloc[0]["code"] == "123"

    @pytest.mark.unit
    def test_read_any_finds_code_column(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Артикул"
        ws["B1"] = "Название"
        ws["A2"] = "WINE123"
        ws["B2"] = "Красное вино"

        excel_file = tmp_path / "test_article.xlsx"
        wb.save(excel_file)

        df = read_any(str(excel_file))

        assert "code" in df.columns
        assert df.iloc[0]["code"] == "WINE123"


# =============================================================================
# Интеграционные тесты с реальной БД (пропускаются по умолчанию)
# =============================================================================


@pytest.mark.integration
@pytest.mark.skipif(
    not RUN_DB_TESTS, reason="Requires local PostgreSQL; enable with RUN_DB_TESTS=1"
)
def test_get_conn_returns_valid_connection(monkeypatch):
    # Настраиваем окружение на локальную БД (docker-compose: 15432 -> 5432)
    monkeypatch.setenv("PGHOST", "localhost")
    monkeypatch.setenv("PGPORT", "15432")
    monkeypatch.setenv("PGUSER", "postgres")
    monkeypatch.setenv("PGPASSWORD", "dev_local_pw")
    monkeypatch.setenv("PGDATABASE", "wine_db")

    conn = get_conn()
    assert conn is not None
    assert isinstance(conn, psycopg2.extensions.connection)

    with conn.cursor() as cur:
        cur.execute("SELECT 1;")
        result = cur.fetchone()
        assert result[0] == 1

    conn.close()


@pytest.mark.integration
@pytest.mark.skipif(
    not RUN_DB_TESTS, reason="Requires local PostgreSQL; enable with RUN_DB_TESTS=1"
)
def test_upsert_records_insert_and_update(monkeypatch):
    """
    Проверяем, что upsert_records:
    - умеет вставлять новую запись;
    - умеет обрабатывать повторный вызов для того же кода на ДРУГУЮ дату,
      не нарушая уникальный индекс (code, effective_from).
    """
    monkeypatch.setenv("PGHOST", "localhost")
    monkeypatch.setenv("PGPORT", "15432")
    monkeypatch.setenv("PGUSER", "postgres")
    monkeypatch.setenv("PGPASSWORD", "dev_local_pw")
    monkeypatch.setenv("PGDATABASE", "wine_db")

    conn = get_conn()
    test_uuid = uuid.uuid4()

    # Уникальный код для этого теста, чтобы не конфликтовать со старыми данными
    code = f"INTTEST_UPSERT_{uuid.uuid4().hex[:8]}"

    # DataFrame вместо списка словарей
    df = pd.DataFrame(
        [
            {
                "code": code,
                "envelope_id": test_uuid,
                "price_rub": 100.0,
                "price_discount": 90.0,
                "price_final_rub": 90.0,
            }
        ]
    )

    today = date.today()

    # Insert
    inserted = upsert_records(df, today)
    assert inserted == 1

    # "Update" того же товара — но с НОВОЙ датой, чтобы не нарушить
    # уникальный индекс (code, effective_from) внутри upsert_price.
    df.loc[0, "price_discount"] = 85.0
    tomorrow = today + timedelta(days=1)
    updated = upsert_records(df, tomorrow)
    assert updated == 1

    conn.close()
