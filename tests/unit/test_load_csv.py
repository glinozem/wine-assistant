import os
import sys
import pandas as pd
from pathlib import Path
from scripts.load_utils import get_conn, _csv_read, _excel_read
from scripts.load_csv import main
from unittest.mock import patch
import pytest


def test_get_conn_returns_connection():
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT 1")
    result = cursor.fetchone()
    assert result == (1,)
    conn.close()


def test_load_csv_parses_file_correctly(tmp_path):
    csv_path = tmp_path / "test.csv"
    csv_path.write_text("Код;Цена\nTEST001;123.45", encoding="utf-8")
    df, *_ = _csv_read(str(csv_path), sep=";")
    assert isinstance(df, pd.DataFrame)
    assert df.shape[0] == 1
    assert "Код" in df.columns


def test_load_excel_parses_basic(tmp_path):
    import openpyxl
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Код"
    ws["A2"] = "TEST001"
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)

    df, *_ = _excel_read(str(xlsx_path), sheet=0, header=0)
    assert isinstance(df, pd.DataFrame)
    assert df.shape[0] == 1
    assert "Код" in df.columns


def test_main_smoke(tmp_path):
    csv_path = tmp_path / "test.csv"
    csv_path.write_text("Код;Цена\nTEST001;123.45", encoding="utf-8")

    with patch("scripts.load_utils._csv_read") as mock_read, \
         patch("scripts.load_utils.upsert_records") as mock_upsert, \
         patch("scripts.load_csv.upsert_records") as mock_main_upsert, \
         patch("scripts.load_csv.check_file_exists") as mock_check_exists, \
         patch("scripts.load_csv.create_envelope") as mock_create_envelope, \
         patch("scripts.load_csv.compute_file_sha256") as mock_hash, \
         patch("scripts.load_csv.update_envelope_status") as mock_update_status, \
         patch("scripts.load_csv.create_price_list_entry") as mock_price_list:

        mock_df = pd.DataFrame({"code": ["TEST001"], "price_rub": [123.45]})
        mock_df.attrs["discount_pct_header"] = None
        mock_read.return_value = (mock_df, ";")
        mock_upsert.return_value = 1
        mock_main_upsert.return_value = 1
        mock_check_exists.return_value = None
        mock_create_envelope.return_value = "00000000-0000-0000-0000-000000000000"
        mock_hash.return_value = "abc123hash"
        mock_price_list.return_value = "plist-id"

        sys.argv = ["load_csv.py", "--csv", str(csv_path)]
        main()


def test_main_fails_on_missing_code(tmp_path):
    csv_path = tmp_path / "test.csv"
    csv_path.write_text("Название;Цена\nUnknown;999", encoding="utf-8")

    with patch("scripts.load_utils._csv_read") as mock_read, \
         patch("scripts.load_csv.check_file_exists") as mock_check_exists, \
         patch("scripts.load_csv.create_envelope") as mock_create_envelope, \
         patch("scripts.load_csv.compute_file_sha256") as mock_hash:

        df = pd.DataFrame({"Название": ["Unknown"], "Цена": [999]})
        df.attrs["discount_pct_header"] = None
        mock_read.return_value = (df, ";")
        mock_check_exists.return_value = None
        mock_create_envelope.return_value = "00000000-0000-0000-0000-000000000000"
        mock_hash.return_value = "abc123hash"

        sys.argv = ["load_csv.py", "--csv", str(csv_path)]
        with pytest.raises(ValueError, match="Не нашли колонку с кодом"):
            main()


def test_main_skips_duplicate_file(tmp_path):
    csv_path = tmp_path / "test.csv"
    csv_path.write_text("Код;Цена\nTEST001;123.45", encoding="utf-8")

    with patch("scripts.load_utils._csv_read") as mock_read, \
         patch("scripts.load_csv.check_file_exists") as mock_check_exists, \
         patch("scripts.load_csv.compute_file_sha256") as mock_hash:

        df = pd.DataFrame({"code": ["TEST001"], "price_rub": [123.45]})
        df.attrs["discount_pct_header"] = None
        mock_read.return_value = (df, ";")
        mock_hash.return_value = "abc123hash"
        mock_check_exists.return_value = {
            "envelope_id": "existing-id",
            "file_name": "existing.csv",
            "status": "imported",
            "upload_timestamp": pd.Timestamp("2024-01-01"),
            "rows_inserted": 1
        }

        sys.argv = ["load_csv.py", "--csv", str(csv_path)]
        main()
