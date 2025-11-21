"""
Unit tests for scripts/date_extraction.py

Цель: Покрытие 80%+ для модуля извлечения дат из файлов.
Issue: #91

Coverage target: 0% → 80%+
"""

from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pytest

from scripts.date_extraction import (
    _parse_date_from_text,
    extract_date_from_excel,
    extract_date_from_filename,
    get_effective_date,
    validate_date,
)

# =============================================================================
# Tests for extract_date_from_excel()
# =============================================================================


@pytest.mark.unit
class TestExtractDateFromExcel:
    """Тесты для функции extract_date_from_excel()"""

    def test_extract_date_from_excel_cell_a1(self, tmp_path):
        """
        Test: Извлечение даты из ячейки A1 в Excel.

        Arrange: Создать Excel файл с датой в A1
        Act: Извлечь дату
        Assert: Дата извлечена корректно
        """
        # Arrange
        test_file = tmp_path / "test_prices.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "Прайс-лист от 20.01.2025"
        workbook.save(test_file)
        workbook.close()

        # Act
        result = extract_date_from_excel(str(test_file), cell="A1")

        # Assert
        assert result is not None
        assert result == date(2025, 1, 20)

    def test_extract_date_from_excel_fallback_to_b1(self, tmp_path):
        """
        Test: Fallback на ячейку B1 если A1 пустая.

        Arrange: Создать Excel с пустой A1 и датой в B1
        Act: Извлечь дату
        Assert: Дата извлечена из B1
        """
        # Arrange
        test_file = tmp_path / "test_fallback.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = None  # Пустая ячейка
        sheet["B1"] = "2025-01-20"
        workbook.save(test_file)
        workbook.close()

        # Act
        result = extract_date_from_excel(str(test_file), cell="A1")

        # Assert
        assert result is not None
        assert result == date(2025, 1, 20)

    def test_extract_date_from_excel_returns_none_if_not_found(self, tmp_path):
        """
        Test: Возвращает None если дата не найдена.

        Arrange: Создать Excel без дат
        Act: Попытаться извлечь дату
        Assert: Возвращает None
        """
        # Arrange
        test_file = tmp_path / "test_no_date.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "Just some text without dates"
        sheet["B1"] = "More text"
        workbook.save(test_file)
        workbook.close()

        # Act
        result = extract_date_from_excel(str(test_file), cell="A1")

        # Assert
        assert result is None

    def test_extract_date_from_excel_multiple_formats(self, tmp_path):
        """
        Test: Поддержка нескольких форматов дат в Excel.

        Arrange: Создать Excel с датами в разных форматах
        Act: Извлечь даты
        Assert: Все форматы распознаны
        """
        # Arrange
        test_file = tmp_path / "test_formats.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        test_formats = {
            "A1": "20.01.2025",  # DD.MM.YYYY
            "A2": "2025-01-20",  # YYYY-MM-DD
            "A3": "20/01/2025",  # DD/MM/YYYY
            "A4": "20-01-2025",  # DD-MM-YYYY
        }

        for cell, value in test_formats.items():
            sheet[cell] = value

        workbook.save(test_file)
        workbook.close()

        # Act & Assert
        for cell in test_formats.keys():
            result = extract_date_from_excel(str(test_file), cell=cell, fallback_cells=[])
            assert result == date(2025, 1, 20), f"Failed to parse date from cell {cell}"

    def test_extract_date_from_excel_handles_corrupted_file(self, tmp_path):
        """
        Test: Обработка поврежденного Excel файла.

        Arrange: Создать некорректный файл
        Act: Попытаться извлечь дату
        Assert: Возвращает None без ошибки
        """
        # Arrange
        test_file = tmp_path / "corrupted.xlsx"
        test_file.write_text("This is not a real Excel file!")

        # Act
        result = extract_date_from_excel(str(test_file))

        # Assert
        assert result is None


# =============================================================================
# Tests for extract_date_from_filename()
# =============================================================================


@pytest.mark.unit
class TestExtractDateFromFilename:
    """Тесты для функции extract_date_from_filename()"""

    def test_extract_date_from_filename_yyyy_mm_dd(self):
        """
        Test: Извлечение даты в формате YYYY_MM_DD.

        Arrange: Имя файла с форматом YYYY_MM_DD
        Act: Извлечь дату
        Assert: Дата извлечена корректно
        """
        # Arrange
        filename = "Price_2025_01_20.xlsx"

        # Act
        result = extract_date_from_filename(filename)

        # Assert
        assert result is not None
        assert result == date(2025, 1, 20)

    def test_extract_date_from_filename_yyyymmdd(self):
        """
        Test: Извлечение даты в формате YYYYMMDD.

        Arrange: Имя файла с форматом YYYYMMDD
        Act: Извлечь дату
        Assert: Дата извлечена корректно
        """
        # Arrange
        filename = "Price_20250120.xlsx"

        # Act
        result = extract_date_from_filename(filename)

        # Assert
        assert result is not None
        assert result == date(2025, 1, 20)

    def test_extract_date_from_filename_dd_mm_yyyy(self):
        """
        Test: Извлечение даты в формате DD.MM.YYYY.

        Arrange: Имя файла с форматом DD.MM.YYYY
        Act: Извлечь дату
        Assert: Дата извлечена корректно
        """
        # Arrange
        filename = "Price_20.01.2025.xlsx"

        # Act
        result = extract_date_from_filename(filename)

        # Assert
        assert result is not None
        assert result == date(2025, 1, 20)

    def test_extract_date_from_filename_yyyy_mm_dd_hyphen(self):
        """
        Test: Извлечение даты с дефисами YYYY-MM-DD.

        Arrange: Имя файла с форматом YYYY-MM-DD
        Act: Извлечь дату
        Assert: Дата извлечена корректно
        """
        # Arrange
        filename = "2025-01-20_price_list.xlsx"

        # Act
        result = extract_date_from_filename(filename)

        # Assert
        assert result is not None
        assert result == date(2025, 1, 20)

    def test_extract_date_from_filename_returns_none_if_not_found(self):
        """
        Test: Возвращает None если дата не найдена.

        Arrange: Имя файла без даты
        Act: Попытаться извлечь дату
        Assert: Возвращает None
        """
        # Arrange
        filename = "price_list_latest.xlsx"

        # Act
        result = extract_date_from_filename(filename)

        # Assert
        assert result is None

    def test_extract_date_from_filename_invalid_date(self):
        """
        Test: Обработка невалидной даты в имени файла.

        Arrange: Имя файла с невалидной датой (например, 32.13.2025)
        Act: Попытаться извлечь дату
        Assert: Возвращает None
        """
        # Arrange
        filename = "Price_2025_13_32.xlsx"  # Невалидная дата

        # Act
        result = extract_date_from_filename(filename)

        # Assert
        assert result is None


# =============================================================================
# Tests for _parse_date_from_text()
# =============================================================================


@pytest.mark.unit
class TestParseDateFromText:
    """Тесты для внутренней функции _parse_date_from_text()"""

    def test_parse_date_from_text_russian_format(self):
        """
        Test: Парсинг даты из русского текста.

        Arrange: Текст с датой в формате DD.MM.YYYY
        Act: Распарсить дату
        Assert: Дата извлечена
        """
        # Arrange
        text = "Прайс-лист от 20.01.2025"

        # Act
        result = _parse_date_from_text(text)

        # Assert
        assert result == date(2025, 1, 20)

    def test_parse_date_from_text_iso_format(self):
        """
        Test: Парсинг даты в ISO формате.

        Arrange: Текст с датой YYYY-MM-DD
        Act: Распарсить дату
        Assert: Дата извлечена
        """
        # Arrange
        text = "Updated on 2025-01-20"

        # Act
        result = _parse_date_from_text(text)

        # Assert
        assert result == date(2025, 1, 20)

    def test_parse_date_from_text_slash_format(self):
        """
        Test: Парсинг даты с слешами DD/MM/YYYY.

        Arrange: Текст с датой DD/MM/YYYY
        Act: Распарсить дату
        Assert: Дата извлечена
        """
        # Arrange
        text = "Effective date: 20/01/2025"

        # Act
        result = _parse_date_from_text(text)

        # Assert
        assert result == date(2025, 1, 20)

    def test_parse_date_from_text_returns_none_if_not_found(self):
        """
        Test: Возвращает None если дата не найдена.

        Arrange: Текст без даты
        Act: Попытаться распарсить
        Assert: Возвращает None
        """
        # Arrange
        text = "This text has no dates in it"

        # Act
        result = _parse_date_from_text(text)

        # Assert
        assert result is None


# =============================================================================
# Tests for validate_date()
# =============================================================================


@pytest.mark.unit
class TestValidateDate:
    """Тесты для функции validate_date()"""

    def test_validate_date_rejects_future_dates(self):
        """
        Test: Валидация отклоняет будущие даты.

        Arrange: Дата в будущем
        Act: Попытаться валидировать
        Assert: Вызывает ValueError
        """
        # Arrange
        future_date = date.today() + timedelta(days=30)

        # Act & Assert
        with pytest.raises(ValueError, match="is in the future"):
            validate_date(future_date)

    def test_validate_date_rejects_old_dates(self):
        """
        Test: Валидация отклоняет даты до 2000 года.

        Arrange: Дата до 2000 года
        Act: Попытаться валидировать
        Assert: Вызывает ValueError
        """
        # Arrange
        old_date = date(1999, 12, 31)

        # Act & Assert
        with pytest.raises(ValueError, match="too old"):
            validate_date(old_date)

    def test_validate_date_accepts_valid_dates(self):
        """
        Test: Валидация принимает корректные даты.

        Arrange: Корректная дата (между 2000 и сегодня)
        Act: Валидировать
        Assert: Возвращает True
        """
        # Arrange
        valid_date = date(2025, 1, 20)

        # Act
        result = validate_date(valid_date)

        # Assert
        assert result is True

    def test_validate_date_accepts_today(self):
        """
        Test: Валидация принимает сегодняшнюю дату.

        Arrange: Сегодняшняя дата
        Act: Валидировать
        Assert: Возвращает True
        """
        # Arrange
        today = date.today()

        # Act
        result = validate_date(today)

        # Assert
        assert result is True


# =============================================================================
# Tests for get_effective_date()
# =============================================================================


@pytest.mark.unit
class TestGetEffectiveDate:
    """Тесты для главной функции get_effective_date()"""

    def test_get_effective_date_priority_asof_override(self, tmp_path):
        """
        Test: Приоритет 1 - asof_override (самый высокий).

        Arrange: Файл с датой в имени + asof_override
        Act: Получить effective date
        Assert: Используется asof_override
        """
        # Arrange
        test_file = tmp_path / "Price_2025_01_10.xlsx"
        test_file.touch()
        override_date = date(2025, 1, 20)

        # Act
        result = get_effective_date(str(test_file), asof_override=override_date)

        # Assert
        assert result == override_date, "asof_override должен иметь наивысший приоритет"

    def test_get_effective_date_priority_excel_cell(self, tmp_path):
        """
        Test: Приоритет 2 - дата из Excel ячейки.

        Arrange: Excel файл с датой в A1
        Act: Получить effective date
        Assert: Используется дата из Excel
        """
        # Arrange
        test_file = tmp_path / "test_excel_priority.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "20.01.2025"
        workbook.save(test_file)
        workbook.close()

        # Act
        result = get_effective_date(str(test_file))

        # Assert
        assert result == date(2025, 1, 20)

    def test_get_effective_date_priority_filename(self, tmp_path):
        """
        Test: Приоритет 3 - дата из имени файла.

        Arrange: Excel без даты + имя файла с датой
        Act: Получить effective date
        Assert: Используется дата из имени
        """
        # Arrange
        test_file = tmp_path / "Price_2025_01_20.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "Some text without date"
        workbook.save(test_file)
        workbook.close()

        # Act
        result = get_effective_date(str(test_file))

        # Assert
        assert result == date(2025, 1, 20)

    def test_get_effective_date_fallback_to_today(self, tmp_path):
        """
        Test: Приоритет 4 - fallback на сегодня.

        Arrange: Файл без дат нигде
        Act: Получить effective date
        Assert: Используется сегодняшняя дата
        """
        # Arrange
        test_file = tmp_path / "price_list_no_date.csv"
        test_file.write_text("product,price\nWine,100")

        # Act
        result = get_effective_date(str(test_file))

        # Assert
        assert result == date.today()

    def test_get_effective_date_fallback_chain(self, tmp_path):
        """
        Test: Полная цепочка fallback (все приоритеты).

        Arrange: Тестировать все уровни fallback
        Act: Получить effective date для каждого случая
        Assert: Правильный приоритет на каждом уровне
        """
        # Test 1: asof_override побеждает всё
        test_file1 = tmp_path / "Price_2025_01_10.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "15.01.2025"
        workbook.save(test_file1)
        workbook.close()

        result1 = get_effective_date(str(test_file1), asof_override=date(2025, 1, 20))
        assert result1 == date(2025, 1, 20), "asof должен побеждать Excel и filename"

        # Test 2: Excel побеждает filename
        result2 = get_effective_date(str(test_file1))
        assert result2 == date(2025, 1, 15), "Excel должен побеждать filename"

        # Test 3: filename если нет Excel
        test_file2 = tmp_path / "Price_2025_01_25.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "No date here"
        workbook.save(test_file2)
        workbook.close()

        result3 = get_effective_date(str(test_file2))
        assert result3 == date(2025, 1, 25), "Должен использовать дату из filename"

        # Test 4: fallback на today если ничего нет
        test_file3 = tmp_path / "nodates.csv"
        test_file3.write_text("data")

        result4 = get_effective_date(str(test_file3))
        assert result4 == date.today(), "Должен fallback на сегодняшнюю дату"

    def test_get_effective_date_validates_extracted_date(self, tmp_path):
        """
        Test: Валидация извлечённой даты.

        Arrange: Файл с будущей датой в имени
        Act: Попытаться получить effective date
        Assert: Вызывает ValueError
        """
        # Arrange
        future_date_str = "Price_2099_12_31.xlsx"
        test_file = tmp_path / future_date_str
        test_file.touch()

        # Act & Assert
        with pytest.raises(ValueError, match="is in the future"):
            get_effective_date(str(test_file))


# =============================================================================
# Integration Tests
# =============================================================================


@pytest.mark.unit
class TestDateExtractionIntegration:
    """
    Интеграционные тесты для полного workflow извлечения дат.
    """

    def test_full_workflow_excel_with_date(self, tmp_path):
        """
        Test: Полный workflow с Excel файлом.

        Arrange: Excel файл с датой в ячейке и в имени
        Act: Извлечь effective date
        Assert: Приоритет работает корректно
        """
        # Arrange
        test_file = tmp_path / "Price_2025_01_15.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "Прайс-лист от 20.01.2025"
        workbook.save(test_file)
        workbook.close()

        # Act
        result = get_effective_date(str(test_file))

        # Assert
        # Excel cell (20.01) должен побеждать filename (15.01)
        assert result == date(2025, 1, 20)

    def test_full_workflow_csv_with_filename_date(self, tmp_path):
        """
        Test: Полный workflow с CSV файлом (нет Excel).

        Arrange: CSV файл с датой только в имени
        Act: Извлечь effective date
        Assert: Используется дата из имени
        """
        # Arrange
        test_file = tmp_path / "price_2025_01_20.csv"
        test_file.write_text("product,price\nWine,100")

        # Act
        result = get_effective_date(str(test_file))

        # Assert
        assert result == date(2025, 1, 20)
