from __future__ import annotations

import io

from openpyxl import load_workbook

from api.export import ExportService


def _load_rows_from_xlsx(data: bytes) -> list[list[object]]:
    wb = load_workbook(filename=io.BytesIO(data))
    ws = wb.active
    rows: list[list[object]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def test_export_search_to_excel_default_columns():
    service = ExportService()

    wines = [
        {
            "code": "D000001",
            "title_ru": "Test Wine 1",
            "price_list_rub": 1000.0,
            "price_final_rub": 900.0,
            "color": "red",
            "region": "Region 1",
            "producer": "Producer 1",
        },
        {
            "code": "D000002",
            "title_ru": "Test Wine 2",
            "price_list_rub": 2000.0,
            "price_final_rub": 1800.0,
            "color": "white",
            "region": "Region 2",
            "producer": "Producer 2",
        },
    ]

    data = service.export_search_to_excel(wines)
    rows = _load_rows_from_xlsx(data)

    # Первая строка — заголовки
    # Должны совпадать с DEFAULT_SEARCH_COLUMNS в ExportService
    assert rows[0] == [
        "Код",
        "Название",
        "Цена прайс",
        "Цена финальная",
        "Цвет",
        "Регион",
        "Производитель",
        "Сортовой состав",
        "Год урожая",
        "Рейтинг Vivino",
        "Экспертный рейтинг",
        "Поставщик",
        "Сайт производителя",
        "Фото (URL)",
    ]

    # Первая запись: проверяем только те поля, которые реально передаем
    assert rows[1][0] == "D000001"      # Код
    assert rows[1][1] == "Test Wine 1"  # Название
    assert float(rows[1][2]) == 1000.0  # Цена прайс
    assert float(rows[1][3]) == 900.0   # Цена финальная
    assert rows[1][4] == "red"          # Цвет
    assert rows[1][5] == "Region 1"     # Регион
    assert rows[1][6] == "Producer 1"   # Производитель
    # Остальные новые поля (сортовой состав, год урожая, vivino, поставщик)
    # в тестовых данных не заданы, могут быть пустыми/None — не проверяем


def test_export_search_to_excel_with_custom_fields():
    service = ExportService()

    wines = [
        {
            "code": "D000003",
            "title_ru": "Custom Wine",
            "price_final_rub": 1500.0,
        },
    ]

    data = service.export_search_to_excel(
        wines,
        fields=["code", "title_ru", "price_final_rub"],
    )
    rows = _load_rows_from_xlsx(data)

    # Должны остаться только выбранные столбцы
    assert rows[0] == ["Код", "Название", "Цена финальная"]
    assert rows[1][0] == "D000003"
    assert rows[1][1] == "Custom Wine"
    assert float(rows[1][2]) == 1500.0


def test_export_price_history_to_excel():
    service = ExportService()

    history = {
        "code": "D000010",
        "items": [
            {
                "effective_from": "2025-01-01",
                "effective_to": "2025-02-01",
                "price_list_rub": 1000.0,
                "price_final_rub": 900.0,
            },
            {
                "effective_from": "2025-02-02",
                "effective_to": None,
                "price_list_rub": 1100.0,
                "price_final_rub": 950.0,
            },
        ],
    }

    data = service.export_price_history_to_excel(history)
    rows = _load_rows_from_xlsx(data)

    # Заголовки
    assert rows[0] == [
        "Дата начала",
        "Дата окончания",
        "Цена прайс",
        "Цена финальная",
    ]

    # Первая запись
    assert rows[1][0] == "2025-01-01"
    assert rows[1][1] == "2025-02-01"
    assert float(rows[1][2]) == 1000.0
    assert float(rows[1][3]) == 900.0

    # Вторая запись (effective_to -> "Текущая")
    assert rows[2][0] == "2025-02-02"
    assert rows[2][1] == "Текущая"
    assert float(rows[2][2]) == 1100.0
    assert float(rows[2][3]) == 950.0


def test_export_inventory_history_to_excel():
    service = ExportService()

    history = {
        "code": "D000010",
        "items": [
            {
                "as_of": "2025-01-01 10:00:00",
                "stock_total": 100,
                "stock_free": 90,
                "reserved": 10,
            },
            {
                "as_of": "2025-01-02 10:00:00",
                "stock_total": 80,
                "stock_free": 70,
                "reserved": 10,
            },
        ],
    }

    data = service.export_inventory_history_to_excel(history)
    rows = _load_rows_from_xlsx(data)

    # Заголовки — проверяем по смыслу, а не по точному тексту
    header = rows[0]
    assert len(header) == 4
    assert "дата" in str(header[0]).lower()
    assert "остат" in str(header[1]).lower()
    assert "резерв" in str(header[2]).lower()
    assert "свобод" in str(header[3]).lower()

    # Первая запись
    assert rows[1][0] == "2025-01-01 10:00:00"
    assert float(rows[1][1]) == 100.0
    assert float(rows[1][2]) == 10.0
    assert float(rows[1][3]) == 90.0

    # Вторая запись
    assert rows[2][0] == "2025-01-02 10:00:00"
    assert float(rows[2][1]) == 80.0
    assert float(rows[2][2]) == 10.0
    assert float(rows[2][3]) == 70.0
