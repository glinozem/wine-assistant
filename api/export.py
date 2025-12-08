from __future__ import annotations

import io
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

SearchItem = Mapping[str, Any]
PriceHistory = Mapping[str, Any]
InventoryHistory = Mapping[str, Any]


# Имена шрифтов по умолчанию (fallback — встроенные Helvetica)
PDF_FONT_REGULAR = "Helvetica"
PDF_FONT_BOLD = "Helvetica-Bold"


def _register_unicode_fonts() -> None:
    """
    Пытаемся зарегистрировать Unicode-шрифт (DejaVu Sans) для поддержки
    кириллицы и символа ₽. Если не получилось — остаёмся на Helvetica.
    """
    global PDF_FONT_REGULAR, PDF_FONT_BOLD

    candidates = [
        # вариант, если когда-нибудь положим ttf рядом с модулем
        Path(__file__).with_name("DejaVuSans.ttf"),
        # стандартный путь в Debian/Ubuntu (fonts-dejavu-core)
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]

    for path in candidates:
        if not path.exists():
            continue
        try:
            pdfmetrics.registerFont(TTFont("DejaVuSans", str(path)))
            # жирный вариант — та же гарнитура, но Bold
            bold_path = path.with_name("DejaVuSans-Bold.ttf")
            if bold_path.exists():
                pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", str(bold_path)))
                PDF_FONT_REGULAR = "DejaVuSans"
                PDF_FONT_BOLD = "DejaVuSans-Bold"
            else:
                # если Bold не нашли — хотя бы обычный
                PDF_FONT_REGULAR = "DejaVuSans"
                PDF_FONT_BOLD = "DejaVuSans"
            break
        except Exception:
            # В dev-окружении без шрифтов просто молча остаёмся на Helvetica,
            # чтобы не ломать тесты.
            continue


_register_unicode_fonts()


@dataclass(frozen=True)
class ExportService:
    """
    Сервис для экспорта данных поиска, карточек и истории цен.

    Пока это чистый сервис без привязки к Flask.
    Flask-эндпоинты будут вызывать его методы.
    """

    # (ключ в dict, заголовок в Excel/PDF)
    DEFAULT_SEARCH_COLUMNS: tuple[tuple[str, str], ...] = (
        ("code", "Код"),
        ("title_ru", "Название"),
        ("price_list_rub", "Цена прайс"),
        ("price_final_rub", "Цена финальная"),
        ("color", "Цвет"),
        ("region", "Регион"),
        ("producer", "Производитель"),

        # новые поля, которые уже есть в products
        ("grapes", "Сортовой состав"),
        ("vintage", "Год урожая"),
        ("vivino_url", "Рейтинг Vivino"),
        ("vivino_rating", "Экспертный рейтинг"),
        ("supplier", "Поставщик"),
        ("producer_site", "Сайт производителя"),
        ("image_url", "Фото (URL)"),
    )

    @staticmethod
    def _fmt_value(value: Any, placeholder: str = "—") -> str:
        if value is None:
            return placeholder
        text = str(value).strip()
        return text or placeholder

    @staticmethod
    def _fmt_qty(value: Any, placeholder: str = "—") -> str:
        if value is None:
            return placeholder
        try:
            num = float(value)
            if num.is_integer():
                return f"{int(num)}"
            return str(num)
        except Exception:
            return str(value)

    @staticmethod
    def _fmt_vivino_score(value: Any, placeholder: str = "—") -> str:
        """
        Красивое форматирование оценки Vivino:
        - если это число, печатаем с одним знаком после запятой;
        - иначе ведём себя как _fmt_value.
        """
        if value is None:
            return placeholder

        text = str(value).strip()
        if not text:
            return placeholder

        try:
            score = float(text.replace(",", "."))
            # типичный диапазон оценок Vivino 0..5
            if 0.0 < score <= 5.0:
                return f"{score:.1f}"
            # если внезапно пришло что-то вроде 96 — считаем, что это не Vivino
            return text
        except (TypeError, ValueError):
            return text

    def export_search_to_excel(
        self,
        wines: Sequence[SearchItem],
        fields: Sequence[str] | None = None,
    ) -> bytes:
        """
        Экспорт результатов поиска в Excel (.xlsx).

        :param wines: последовательность dict-подобных объектов с данными по винам
        :param fields: список ключей, которые нужно включить (настраиваемые поля).
                       Если None — используется DEFAULT_SEARCH_COLUMNS целиком.
        """
        columns = [
            col
            for col in self.DEFAULT_SEARCH_COLUMNS
            if fields is None or col[0] in fields
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Wine Search Results"

        # Заголовки
        headers = [header for _, header in columns]
        ws.append(headers)

        header_fill = PatternFill(
            start_color="4F81BD",
            end_color="4F81BD",
            fill_type="solid",
        )
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Данные
        for wine in wines:
            row = [wine.get(field, "") for field, _ in columns]
            ws.append(row)

        # Автоширина колонок
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value is None:
                    continue
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            ws.column_dimensions[column_letter].width = max_length + 2

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_search_to_pdf(
        self,
        wines: Sequence[SearchItem],
        max_title_len: int = 30,
    ) -> bytes:
        """
        Экспорт результатов поиска в PDF (таблица).

        Формат простой, оптимизирован под печать/быстрый просмотр.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements: list[Any] = []

        styles = getSampleStyleSheet()
        title_style = styles["Title"]
        title_style.fontName = PDF_FONT_BOLD
        title = Paragraph("<b>Wine Search Results</b>", title_style)
        elements.append(title)

        data: list[list[str]] = [
            [
                "Код",
                "Название",
                "Цена",
                "Цвет",
                "Регион",
                "Сорт",
                "Год",
                "Рейтинг Vivino",
                "Экспертный рейтинг",
                "Сайт производителя",
                "Фото (URL)",
            ],
        ]

        for wine in wines:
            title_ru = (wine.get("title_ru") or "")[:max_title_len]
            price_final = wine.get("price_final_rub") or 0
            grapes = wine.get("grapes") or ""
            vintage = wine.get("vintage") or ""

            vivino_score = wine.get("vivino_url") or ""
            expert_score = wine.get("vivino_rating") or ""
            producer_site = (wine.get("producer_site") or "").strip()
            image_url = wine.get("image_url") or ""

            data.append(
                [
                    str(wine.get("code", "")),
                    title_ru,
                    f"{price_final:.0f} ₽",
                    str(wine.get("color", "") or ""),
                    str(wine.get("region", "") or ""),
                    str(grapes),
                    str(vintage),
                    str(vivino_score),
                    str(expert_score),
                    producer_site,
                    image_url,
                ]
            )

        table = Table(data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("FONTNAME", (0, 1), (-1, -1), PDF_FONT_REGULAR),
                ]
            )
        )

        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        return buffer.getvalue()

    def export_wine_card_to_pdf(self, wine: Mapping[str, Any]) -> bytes:
        """
        Экспорт одной карточки товара в PDF.
        """
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        # Заголовок: title_ru или name
        title = wine.get("title_ru") or wine.get("name") or ""
        title = str(title)
        if title:
            c.setFont(PDF_FONT_BOLD, 16)
            c.drawString(2 * cm, height - 3 * cm, title)

        y = height - 4.5 * cm

        # Страна под названием, если есть
        country = wine.get("country")
        if country:
            c.setFont(PDF_FONT_REGULAR, 12)
            c.drawString(2 * cm, y, str(country))
            y -= 1.0 * cm

        c.setFont(PDF_FONT_REGULAR, 11)

        # Если reserved не пришёл — считаем его здесь
        reserved = wine.get("reserved")
        if reserved is None:
            stock_total = wine.get("stock_total")
            stock_free = wine.get("stock_free")
            try:
                if stock_total is not None and stock_free is not None:
                    reserved = float(stock_total) - float(stock_free)
            except Exception:
                reserved = None

        # По данным текущих прайсов колонка "Vivino" содержит именно оценку,
        # а не ссылку, поэтому используем vivino_url как источник рейтинга.
        vivino_score_raw = wine.get("vivino_url") or wine.get("vivino_rating")
        vivino_score = self._fmt_vivino_score(vivino_score_raw)

        fields: list[tuple[str, str]] = [
            ("Код товара", self._fmt_value(wine.get("code"))),
            ("Производитель", self._fmt_value(wine.get("producer"))),
            ("Фото (URL)", self._fmt_value(wine.get("image_url"))),
            ("Регион", self._fmt_value(wine.get("region"))),
            ("Цвет", self._fmt_value(wine.get("color"))),
            ("Стиль", self._fmt_value(wine.get("style"))),
            ("Сорт винограда", self._fmt_value(wine.get("grapes"))),
            ("Год урожая", self._fmt_value(wine.get("vintage"))),
            ("Поставщик", self._fmt_value(wine.get("supplier"))),
            ("Сайт производителя", self._fmt_value(wine.get("producer_site"))),
            ("Рейтинг Vivino", vivino_score),
            ("Экспертный рейтинг", self._fmt_value(wine.get("vivino_rating"))),
            ("", ""),
            ("Цена прайс", self._fmt_price(wine.get("price_list_rub"))),
            ("Цена финальная", self._fmt_price(wine.get("price_final_rub"))),
            ("", ""),
            ("Остаток (всего)", self._fmt_qty(wine.get("stock_total"))),
            ("Зарезервировано", self._fmt_qty(reserved)),
            ("Свободно", self._fmt_qty(wine.get("stock_free"))),
        ]


        for label, value in fields:
            if y < 2 * cm:
                c.showPage()
                y = height - 3 * cm
                c.setFont(PDF_FONT_REGULAR, 11)

            if not label:
                y -= 0.5 * cm
                continue

            c.setFont(PDF_FONT_BOLD, 11)
            c.drawString(2 * cm, y, f"{label}:")
            c.setFont(PDF_FONT_REGULAR, 11)
            c.drawString(8 * cm, y, value)
            y -= 0.7 * cm

        c.showPage()
        c.save()
        buffer.seek(0)
        return buffer.getvalue()

    def export_price_history_to_excel(
        self,
        history: PriceHistory,
    ) -> bytes:
        """
        Экспорт истории цен в Excel.

        Ожидаемый формат history:
        {
            "code": "D011283",
            "items": [
                {
                    "effective_from": "...",
                    "effective_to": "... или None",
                    "price_list_rub": 123.45,
                    "price_final_rub": 100.00,
                },
                ...
            ]
        }
        """
        wb = Workbook()
        ws = wb.active

        code = history.get("code") or ""
        ws.title = f"Price History {code}"

        ws.append(["Дата начала", "Дата окончания", "Цена прайс", "Цена финальная"])

        for item in history.get("items", []):
            effective_from = item.get("effective_from")
            effective_to = item.get("effective_to") or "Текущая"
            price_list = item.get("price_list_rub")
            price_final = item.get("price_final_rub")

            ws.append(
                [
                    effective_from,
                    effective_to,
                    price_list,
                    price_final,
                ]
            )

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


    def export_inventory_history_to_excel(
        self,
        history: InventoryHistory,
    ) -> bytes:
        """
        Экспорт истории остатков в Excel.

        Ожидаемый формат history:
        {
            "code": "D011283",
            "items": [
                {
                    "as_of": "2025-10-27 12:00:00",
                    "stock_total": 12502,
                    "stock_free": 12334,
                    "reserved": 0,
                },
                ...
            ],
            "total": 2,
            "limit": 50,
            "offset": 0,
        }
        """
        wb = Workbook()
        ws = wb.active

        code = history.get("code") or ""
        ws.title = f"Inventory History {code}"

        # Заголовок
        ws.append(
            [
                "Дата (as_of)",
                "Остаток (всего)",
                "Зарезервировано",
                "Свободно",
            ]
        )

        for item in history.get("items", []):
            stock_total = item.get("stock_total")
            stock_free = item.get("stock_free")
            reserved = item.get("reserved")

            # Пытаемся аккуратно пересчитать reserved = stock_total - stock_free
            try:
                if stock_total is not None and stock_free is not None:
                    reserved_calc = float(stock_total) - float(stock_free)

                    # Если результат целый — выводим как int, чтобы в Excel было красиво
                    if reserved_calc.is_integer():
                        reserved = int(reserved_calc)
                    else:
                        reserved = reserved_calc
            except Exception:
                # Если расчёт не удался, оставляем reserved как есть
                pass

            ws.append(
                [
                    item.get("as_of"),
                    stock_total,
                    reserved,
                    stock_free,
                ]
            )

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()



    @staticmethod
    def _fmt_price(value: Any) -> str:
        if value is None:
            return "N/A"
        try:
            return f"{float(value):.0f} ₽"
        except (TypeError, ValueError):
            return str(value)
