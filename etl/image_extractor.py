from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Dict, Optional

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.worksheet import Worksheet


def _make_safe_filename(code: str, ext: str) -> str:
    """
    Делает безопасное имя файла из кода:
    - убираем слэши, обратные слэши и прочие спецсимволы,
    - оставляем буквы/цифры/подчёркивания/дефисы,
    - всё остальное заменяем на "_".
    """
    if code is None:
        code = ""
    code_str = str(code)

    # Заменяем явные разделители путей
    code_str = code_str.replace("/", "_").replace("\\", "_")

    # Оставляем только [0-9a-zA-Z_-.] и кириллицу, остальное -> "_"
    safe = []
    for ch in code_str:
        if ch.isalnum() or ch in ("_", "-", "."):
            safe.append(ch)
        else:
            safe.append("_")
    safe_code = "".join(safe).strip("_")

    if not safe_code:
        safe_code = "image"

    if not ext.startswith("."):
        ext = "." + ext

    return f"{safe_code}{ext}"


def _find_code_column(ws: Worksheet, header_row: int, code_header: str = "Код") -> int | None:
    """
    Находит индекс колонки с заголовком 'Код' в заданной строке header_row (1-based).
    Возвращает 0-based индекс колонки или None, если не найдено.
    """
    for idx, cell in enumerate(ws[header_row]):
        value = str(cell.value).strip() if cell.value is not None else ""
        if value == code_header:
            return idx
    return None


def _build_row_to_code_map(ws: Worksheet, header_row: int, code_col_idx: int) -> Dict[int, str]:
    """
    Строит мапу: номер строки (1-based) -> SKU code.

    header_row — строка с заголовками (1-based).
    """
    row_to_code: Dict[int, str] = {}
    # Данные начинаются со строки header_row + 1
    for row in ws.iter_rows(min_row=header_row + 1):
        cell = row[code_col_idx]
        if cell.value is None:
            continue
        code = str(cell.value).strip()
        if not code:
            continue
        row_to_code[cell.row] = code
    return row_to_code


def _ensure_output_dir(base_dir: Path) -> Path:
    base_dir.mkdir(parents=True, exist_ok=True)
    return base_dir


def extract_images_from_excel(
    excel_path: str | Path,
    header_row_zero_based: Optional[int] = 3,
    code_header: str = "Код",
    output_dir: str | Path | None = None,
    image_base_url: str | None = None,
) -> Dict[str, str]:
    """
    Извлекает встроенные картинки из Excel-файла и привязывает их к SKU по строке.

    Args:
        excel_path: путь к исходному .xlsx прайсу.
        header_row_zero_based: номер строки заголовков в терминах pandas (0-based),
                               у тебя сейчас в логах это 3 -> строка 4 в Excel.
        code_header: текст заголовка колонки с кодом товара (обычно 'Код').
        output_dir: директория, куда складывать файлы картинок.
                    Если не указана, берется WINE_IMAGE_DIR или 'static/images'.
        image_base_url: базовый URL для формирования image_url.
                        Если не указан, берется WINE_IMAGE_BASE_URL.

    Returns:
        dict: {sku_code: image_url}
    """
    excel_path = Path(excel_path)
    if not excel_path.is_file():
        return {}

    # Если вдруг кто-то передал None, откатываемся на дефолт 3
    if header_row_zero_based is None:
        header_row_zero_based = 3

    # 1. Настраиваем директорию и базовый URL
    if output_dir is None:
        output_dir = os.getenv("WINE_IMAGE_DIR", "static/images")
    output_dir_path = _ensure_output_dir(Path(output_dir))

    if image_base_url is None:
        image_base_url = os.getenv("WINE_IMAGE_BASE_URL", "").rstrip("/")

    # 2. Открываем книгу
    wb = load_workbook(filename=str(excel_path), data_only=True)
    ws = wb.active  # у тебя прайс всегда на первом листе

    # 3. Находим колонку с кодом
    header_row_excel = header_row_zero_based + 1  # pandas header=3 -> Excel row 4 (1-based)
    code_col_idx = _find_code_column(ws, header_row_excel, code_header=code_header)
    if code_col_idx is None:
        # Не нашли колонку с кодом — ничего не делаем
        return {}

    # 4. Строим мапу "строка -> код"
    row_to_code = _build_row_to_code_map(ws, header_row_excel, code_col_idx)
    if not row_to_code:
        return {}

    # 5. Если картинок нет — просто выходим
    #   (openpyxl хранит их в приватном списке _images)
    images = getattr(ws, "_images", []) or []
    if not images:
        return {}

    code_to_url: Dict[str, str] = {}

    for img in images:
        if not isinstance(img, XLImage):
            continue

        # anchor._from.row / .col — 0-based индексы
        try:
            anchor = img.anchor._from  # type: ignore[attr-defined]
            row_index = anchor.row + 1  # делаем 1-based, как в ws
        except Exception:
            # Если вдруг не смогли распарсить anchor — пропускаем
            continue

        code = row_to_code.get(row_index)
        if not code:
            # Картинка не привязана к строке с кодом товара — пропускаем
            continue

        # Не перезаписываем, если у кода уже есть картинка
        if code in code_to_url:
            continue

        # Определяем расширение файла
        ext = (img.format or "png").lower()
        if not ext.startswith("."):
            ext = "." + ext

        # Делаем безопасное имя файла
        filename = _make_safe_filename(code, ext)
        file_path = output_dir_path / filename

        # 6. Сохраняем бинарные данные картинки
        try:
            data = img._data()
            with open(file_path, "wb") as f:
                f.write(data)
        except Exception as e:
            # Логируем и продолжаем, не роняем весь процесс
            print(f"[images] Failed to write image for code={code!r}: {e}")
            continue

        # 7. Формируем URL
        if image_base_url:
            url = f"{image_base_url}/{filename}"
        else:
            # fallback: относительный путь
            url = str(file_path.as_posix())

        code_to_url[code] = url

    return code_to_url
